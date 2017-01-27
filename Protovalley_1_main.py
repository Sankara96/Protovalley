from PySide import QtGui,QtCore
from Prototype_1 import Ui_Proto
import sys
from openpyxl.cell import column_index_from_string,get_column_letter
import openpyxl
import os

class ProtoApp(QtGui.QMainWindow):
    def __init__(self):
        QtGui.QMainWindow.__init__(self)
        self.ui = Ui_Proto()
        self.ui.setupUi(self)
        self.order_dict = {}
        self.excel = openpyxl.load_workbook('order-details.xlsx')
        self.excel_sheet = self.excel.get_sheet_by_name('Sheet1')
        self.excel_sheet_col_num = self.excel_sheet.max_column
        self.excel_sheet_row_num = self.excel_sheet.max_row
        self.order_dict["design"] = 0
        self.order_dict["dlp"] = 0
        self.order_dict["fdm"] = 0
        self.order_dict["casting"] = 0
        self.reference = str(self.excel_sheet['F1'].value)
        self.actual = 'PV'+str(self.excel_sheet['F1'].value).zfill(4)
        self.ui.reference_number_value.setText(self.actual)
        self.ui.order_group.setEnabled(False)
        self.ui.design.stateChanged.connect(self.check)
        self.ui.dlp.stateChanged.connect(self.check)
        self.ui.fdm.stateChanged.connect(self.check)
        self.ui.casting.stateChanged.connect(self.check)
        self.ui.done_btn.clicked.connect(self.add_details)
        self.ui.new_order.stateChanged.connect(self.activate)
        self.ui.add_btn.clicked.connect(self.add)

    def add(self):
        browse_path = "C:"
        file_path = QtGui.QFileDialog.getOpenFileName(QtGui.QFileDialog(),
                                                      'Select File', browse_path)
        self.file_name = file_path[0]
        self.file_widget  = QtGui.QTreeWidgetItem(self.ui.statusbox)
        self.file_widget.setText(0,self.file_name)

    def activate(self):
        if self.ui.new_order.isChecked() == True:
            self.ui.order_group.setEnabled(True)
        else:
            self.ui.order_group.setEnabled(False)
    def add_details(self):
        self.excel_sheet['A'+str(self.excel_sheet_row_num+1)].value = 'PV'+str(self.reference).zfill(4)
        self.excel_sheet['B'+str(self.excel_sheet_row_num+1)].value = self.order_dict['design']
        self.excel_sheet['C' + str(self.excel_sheet_row_num + 1)].value = self.order_dict['dlp']
        self.excel_sheet['D' + str(self.excel_sheet_row_num + 1)].value = self.order_dict['fdm']
        self.excel_sheet['E'+str(self.excel_sheet_row_num+1)].value = self.order_dict['casting']
        self.excel_sheet['F1'] = str(int(self.reference)+1)
        self.excel.save('order-details.xlsx')
        root = self.ui.statusbox.invisibleRootItem()
        child = root.childCount()
        if child>0:
            newpath = "C:/Users/ADMIN/PycharmProjects/Protovalley/"+str(self.actual)
            if not os.path.exists(newpath):
                os.makedirs(newpath)

            for i in range(child):
                item = root.child(i)
                elements = item.text(0)
                temp = ''
                for i in elements[::-1]:
                    if i =='/' or i =='\\':
                        break
                    elif i!='/' or i!= r'\'':
                        temp = ''.join((i,temp))
                os.rename(elements,newpath+'/'+str(temp))
        else:
            pass
        exit()





    def check(self):

        if self.ui.design.isChecked() == True:
            self.order_dict['design'] = 1
        else:
            self.order_dict['design'] = 'x'

        if self.ui.fdm.isChecked() == True:
            self.order_dict['fdm'] = 1
        else:
            self.order_dict['fdm'] = 'x'

        if self.ui.dlp.isChecked() == True:
            self.order_dict['dlp'] = 1
        else:
            self.order_dict['dlp'] = 'x'

        if self.ui.casting.isChecked() == True:
            self.order_dict['casting'] = 1
        else:
            self.order_dict['casting'] = 'x'


    def closeEvent(self, event):
        reply = QtGui.QMessageBox.question(QtGui.QMessageBox(), 'Message',
                                           "Are you sure to quit?", QtGui.QMessageBox.Yes | QtGui.QMessageBox.No,
                                           QtGui.QMessageBox.No)
        if reply == QtGui.QMessageBox.Yes:
            self.exit = True
            event.accept()
        else:
            event.ignore()
if __name__ == '__main__':

    Protovalley_app = QtGui.QApplication(sys.argv)
    Proto_window = ProtoApp()
    Proto_window.show()
    status = Protovalley_app.exec_()
